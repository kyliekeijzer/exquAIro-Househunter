import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_metrics_to_excel(models: dict, filename: str = "model_comparison"):
    """
    Formats model performance metrics into a comparison table and saves to Excel.

    Args:
        models: dict of {model_name: model_output} where model_output is the
                dict returned by perform_linear_regression or perform_logistic_regression.
                Example: {"Linear": linear_model, "Logistic": logistic_model}
        filename: output filename without extension (default: "model_comparison")
    """
    rows = []
    for model_name, model_output in models.items():
        for split in ["train", "test"]:
            metrics = model_output.get(f"{split}_metrics", {})
            row = {
                "Model": model_name,
                "Split": split.capitalize(),
                **{k: v for k, v in metrics.items() if k != "confusion_matrix"}
            }
            rows.append(row)

    df = pd.DataFrame(rows)

    # Drop columns that are entirely empty (e.g. R² for logistic)
    df.dropna(axis=1, how="all", inplace=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Results"

    # Styles
    header_font    = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill    = PatternFill("solid", start_color="2F5496")
    model_font     = Font(name="Arial", bold=True)
    normal_font    = Font(name="Arial")
    center_align   = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = model_font if col_idx == 1 else normal_font
            cell.alignment = center_align
            cell.border    = thin_border

            # Format numeric cells to 4 decimal places
            if isinstance(value, float):
                cell.number_format = "0.0000"

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    "solid", start_color="DCE6F1"
                )

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    filepath = f"{filename}.xlsx"
    wb.save(filepath)
    print(f"Saved to {filepath}")
    return filepath